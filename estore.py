from openpyxl import load_workbook as lw
from time import strftime as st, sleep
from tkinter import messagebox as msg
import sys, os

dia = st("%Y-%m-%d 00:00:00")
data = st("%d/%m/%Y")
mm = st('%Y%m')
ly = int(st('%Y')) - 1
ly = st(f'{ly}%m')
try:
    pln = lw(f"metas//metas.xlsx")
except:
    msg.showinfo('Automação', 'Metas não atualizadas')
    
    
ws = pln.active 
soma = 0
total = 0

class es():
    def iniciar(self):
        def L(lj):
            global soma
            soma = 0
            linhas = 6000
            for i in range(4, linhas):
                diaP = ws[f"A{i}"].value
                diaP = str(diaP)
                loja = ws[f"B{i}"].value
                dco = ws[f"C{i}"].value
                valor = ws[f"D{i}"].value
                # Retira celulas vazias 
                if diaP != None:
                    if diaP == dia:
                        if loja == f'L{lj}':
                            soma += valor
            soma = soma * 0.01
            soma = round(soma)

        L("025")
        l25 = soma

        L("051")
        l51 = soma

        L("054")
        l54 = soma

        L("056")
        l56 = soma

        L("057")
        l57 = soma

        L("059")
        l59 = soma

        L("061")
        l61 = soma

        L("062")
        l62 = soma

        L("194")
        l194 = soma

        L("201")
        l201 = soma

        L("287")
        l287 = soma

        L("306")
        l306 = soma

        L("313")
        l313 = soma

        L("316")
        l316 = soma

        L("326")
        l326 = soma

        L("391")
        l391 = soma

        L("393")
        l393 = soma

        total = l25 + l51 + l54 + l56 + l57 +l59 +l61 + l62 + l194 + l201 + l287 + l306+l313 + l316 + l326 + l391 + l393
        '''
        😝 🗓
        '''

        text = f'''
*METAS DIÁRIAS E-STORE CTB*
*{data}*

_*Meta//Realizado*_

L025: R${l25},00 //
L051: R${l51},00 //
L054: R${l54},00 //
L056: R${l56},00 //
L057: R${l57},00 //
L059: R${l59},00 //
L061: R${l61},00 //
L062: R${l62},00 //
L194: R${l194},00 //
L201: R${l201},00 //
L287: R${l287},00 //
L306: R${l306},00 //
L313: R${l313},00 //
L316: R${l316},00 //
L326: R${l326},00 //
L391: R${l391},00 //
L393: R${l393},00 //


*TOTAL: R$ {total},00*

*REALIZADO: R$ 0*


*FEZ A VENDA ATUALIZA!* '''
        a = open('txt\\estore.txt', 'w')
        a.write(text)
        a.close()
        os.system('cd txt && estore.txt')

