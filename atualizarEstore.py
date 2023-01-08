from openpyxl import load_workbook as lw

class esMes():
    def __init__(self, lj):
        meta = lw('metas\\metas.xlsx')
        ws = meta.active
        self.soma = 0
        for i in range(4, 6000):
            loja = ws[f'B{i}'].value
            if loja == lj:
                valor = ws[f'D{i}'].value
                self.soma += valor
        self.final = self.soma * 0.01
class init():
    def __init__(self) -> None:
        plne = lw('pln\\estore.xlsx')
        ws2 = plne.active
        fls = [
            'L025',
            'L051',
            'L054',
            'L056',
            'L057',
            'L059',
            'L061',
            'L062',
            'L194',
            'L201',
            'L287',
            'L306',
            'L313',
            'L316',
            'L326',
            'L391',
            'L393'
            ]
        cont = 8
        for i in fls:
            e = esMes(i)
            ws2[f'C{cont}'] = e.final
            cont += 1
        plne.save('pln\\estore.xlsx')
        plne.close()
