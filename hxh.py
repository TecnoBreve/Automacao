from openpyxl import load_workbook
from metas import Metas
import time, os

#Carregando Data
dia = time.strftime("DATA %d/%m/%Y")
#Carregando e Ativando xlsx
planilha = load_workbook(r'pln//horaxhora.xlsx')
ws = planilha.active


class DCO():
    def __init__(self):
        #Valor Generico
        self.generico = float()
        #Departamentos
        self.infantil = float()
        self.fem = float()
        self.relo = float()
        self.biju = float()
        self.oculos = float()
        self.basket = float()
        self.masc = float()
        self.mdc = float()
        self.cba = float()
        self.tecno = float()
        self.blz = float()
        self.metas()
    def metas(self):
        m = Metas()
        ws['C7'] = m.blz
        ws['C8'] = m.cba
        ws['C9'] = m.tec
        ws['C10'] = m.fem
        ws['C11'] = m.inf
        ws['C12'] = m.masc
        ws['C13'] = m.mdc
        ws['C14'] = m.biju
        ws['C15'] = m.basket
        ws['C16'] = m.rel
        #planilha.save('horaxhora.xlsx')

    def ler(self, tdco):
        #Ler dados
        with open(r'txt//texto.txt') as t:
            self.dados = t.readlines()

        #Lendo linha por DCO
        for linha in self.dados:
            if tdco in linha:
                #separador e localizacao de valor
                value = linha.split()
                c = len(value)
                if c == 10:
                    value = value[5]
                if c == 11:
                    value = value[6]
                if c == 12:
                    value = value[7]
                if c == 13:
                    value = value[8]
                if c == 14:
                    value = value[5]
                if c == 15:
                    value = value[6]
                if c == 16:
                    value = value[7]
                if c == 17:
                    value = value[8]
                #retirar dezena de milhar
                if value.count(".") >= 1:
                    value = value.replace(".", "")
                #Converter STR em Float
                value = value.replace(",", ".")
                value = float(value)
                #Inserir ao valor generico
                self.generico += (value)
            else:
                self.generico += 0
                
def mainHxh():
    d = DCO()
    def inf():
        d.generico = float()
        d.ler('100-')
        d.ler('101-')
        d.ler('102-')
        d.ler('105-')
        d.ler('106-')
        d.ler('110-')
        d.ler('115-')
        d.ler('120-')
        d.ler('125-')
        d.ler('130-')
        d.ler('131-')
        d.ler('135-')
        d.ler('142-')
        d.ler('145-')
        d.ler('148-')
        d.ler('150-')
        d.ler('151-')
        d.ler('165-')
        d.ler('175-')
        d.ler('195-')
        d.ler('196-')
        d.infantil = d.generico

    def biju():
        d.generico = float()
        d.ler('560-')
        d.ler('565-')
        d.biju = d.generico

    def bskt():
        d.generico = float()
        d.ler('572-')
        d.basket = d.generico

    def rel():
        d.generico = float()
        d.ler('550-')
        d.relo = d.generico

    def fem():
        d.generico = float()
        d.ler('180-')
        d.ler('200-')
        d.ler('204-')
        d.ler('205-')
        d.ler('206-')
        d.ler('208-')
        d.ler('210-')
        d.ler('215-')
        d.ler('216-')
        d.ler('217-')
        d.ler('220-')
        d.ler('225-')
        d.ler('226-')
        d.ler('227-')
        d.ler('230-')
        d.ler('235-')
        d.ler('240-')
        d.ler('245-')
        d.ler('250-')
        d.ler('255-')
        d.ler('260-')
        d.ler('261-')
        d.ler('265-')
        d.ler('270-')
        d.ler('275-')
        d.ler('290-')
        d.ler('295-')
        d.ler('296-')
        d.fem = d.generico

    def masc():
        d.generico = float()
        d.ler('300-')
        d.ler('301-')
        d.ler('305-')
        d.ler('306-')
        d.ler('310-')
        d.ler('314-')
        d.ler('315-')
        d.ler('320-')
        d.ler('325-')
        d.ler('330-')
        d.ler('331-')
        d.ler('340-')
        d.ler('350-')
        d.ler('360-')
        d.ler('365-')
        d.ler('370-')
        d.ler('375-')
        d.ler('380-')
        d.ler('390-')
        d.ler('395-')
        d.masc = d.generico

    def mdc():
        d.generico = float()
        d.ler('400-')
        d.ler('405-')
        d.ler('410-')
        d.ler('415-')
        d.ler('420-')
        d.ler('425-')
        d.ler('430-')
        d.ler('435-')
        d.ler('440-')
        d.ler('445-')
        d.mdc = d.generico

    def cba():
        d.generico = float()
        d.ler('500-')
        d.ler('505-')
        d.ler('510-')
        d.ler('515-')
        d.ler('520-')
        d.ler('540-')
        d.ler('570-')
        d.ler('595-')
        d.cba = d.generico

    def tec():
        d.generico = float()
        d.ler('530-E')
        d.ler('531-')
        d.ler('532-')
        d.ler('533-')
        d.ler('551-')
        d.ler('552-')
        d.tecno = d.generico

    def blz():
        d.generico = float()
        d.ler('280-')
        d.ler('281-')
        d.ler('282-')
        d.ler('283-')
        d.ler('284-')
        d.ler('285-')
        d.ler('289-')
        d.blz = d.generico

    def inserir():
        ws["A4"] = dia
        ws["E11"] = d.infantil
        ws["E14"] = d.biju
        ws["E15"] = d.basket
        ws["E16"] = d.relo
        ws["E10"] = d.fem
        ws['E7'] = d.blz
        ws['E8'] = d.cba
        ws['E9'] = d.tecno
        ws['E12'] = d.masc
        ws['E13'] = d.mdc
    blz()
    cba()
    tec()
    fem()
    inf()
    masc()
    mdc()
    biju()
    bskt()
    rel()
    inserir()
    planilha.save(r"pln//horaxhora.xlsx")
    planilha.close()
    os.system('cd pln && horaxhora.xlsx')
