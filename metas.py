from openpyxl import load_workbook as lw
from time import strftime as st

dia = st('%Y-%m-%d 00:00:00')
mes = st('%Y%m')

  
class Metas():
    def __init__(self):
        self.soma = float()
        self.basket = float()
        self.blz = float()
        self.cba = float()
        self.tec = float()
        self.fem = float()
        self.inf = float()
        self.mas = float()
        self.mdc = float()
        self.biju = float()
        self.rel = float()
        try:
            self.planilha = lw(f"metas\\metas.xlsx")
            self.ws = self.planilha.active
            self.grupo()
        except:
            pass

    def grupo(self):
        for i in range(4, 5000):
            diaP = self.ws[f'A{i}'].value
            valor = self.ws[f'D{i}'].value
            fl = self.ws[f'B{i}'].value
            grupo = self.ws[f"C{i}"].value
            diaP = str(diaP)
            if diaP == dia:
                if fl == self.fl:
                    if grupo == 'Basket':
                        self.basket = valor
                        self.soma += valor
                    if grupo == 'Perfumaria':
                        self.blz = valor
                        self.soma += valor
                    if grupo == 'Calçados':
                        self.cba = valor
                        self.soma += valor
                    if grupo == 'Eletrônicos':
                        self.tec = valor
                        self.soma += valor
                    if grupo == 'Feminino':
                        self.fem = valor
                        self.soma += valor
                    if grupo == 'Infanto-Juvenil':
                        self.inf = valor
                        self.soma += valor
                    if grupo == 'Masculino':
                        self.masc = valor
                        self.soma += valor
                    if grupo == 'Moda Casa':
                        self.mdc = valor
                        self.soma += valor
                    if grupo == 'Óculos e Bijuteria':
                        self.biju = valor
                        self.soma += valor
                    if grupo == 'Relógios':
                        self.rel = valor
                        self.soma += valor

'''m = Metas
m.fl = 'L201'
m()
print(m().soma)'''
