from openpyxl import load_workbook as lw
import os

class VJ():
    def __init__(self):
        self.arquivo = open(r"txt//vj.txt")
        self.planilha = lw(r"pln//vj.xlsx")
        self.ws = self.planilha.active
        
        # Vars
        self.valorE = 0
        self.soma = float()
        
    def ler(self):
        for i in range(1,7):
            b = self.arquivo.readline()
            b = b.replace("\n","")
            b = b.replace(" ","")
            b = b.replace(".","")
            b = b.replace(",",".")
        self.valor = b
        self.valor = float(self.valor)
        
    def entrada(self):
        for i in range(1,9):
            b = self.arquivo.readline()
            b = b.replace("\n","")
            b = b.replace(" ","")
            b = b.replace(".","")
            b = b.replace(",",".")
        self.valorE = b
        self.valorE = float(self.valorE)
        
    def operador(self, a):
        for i in range(1, a):
            self.op = self.arquivo.readline()
        self.op = self.op.replace("\n","")
        for i in range(4, 51):
            ope = self.ws[f'G{i}'].value
            if ope == self.op:
                self.nome = self.ws['A{i}']
                break

    def ins(self):
        for conti in range(4,100):
            self.opp = self.ws[f"G{conti}"].value
            if self.opp != None:
                if int(self.opp) == int(self.op):
                    oldv = self.ws[f"C{conti}"].value
                    if oldv == None:
                        oldv = 0
                    self.valor = self.valor - self.valorE
                    newv = oldv + self.valor
                    self.ws[f"C{conti}"] = newv
    def code(self):
        # INIT CODE               
        for i in range(4, 27):
            self.ws[f'C{i}'] = 0

        for cont in range(1, 3000):
            a = self.arquivo.readline()
            a = a.replace("\n","")
            a = a.split()
            
            # PDV
            for j in range(40, 400 ):
                    if str(j) in a:
                        self.ler()
                        self.operador(9)
                        self.entrada()
                        self.soma += self.valor
                        self.ins()
            # Mobile
            for j in range(500, 1000 ):
                    if str(j) in a:
                        self.ler()
                        self.operador(7)
                        self.entrada()
                        self.soma += self.valor
                        self.ins()
                        
        self.planilha.save(r"pln//vj.xlsx")
        self.planilha.close()
        os.system('cd pln && vj.xlsx')
