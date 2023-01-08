from openpyxl import load_workbook as lw
import os 

class PA():
    def __init__(self):
        self.planilha = lw(r"pln//parcial.xlsx")
        self.ws = self.planilha.active
        ws = self.ws
        for i in range(4, 24):
            ws[f'C{i}'] = 0
            ws[f'E{i}'] = 0
        for i in range(25, 29):
            ws[f'C{i}'] = 0
            ws[f'E{i}'] = 0
        
    def posP(self):
        self.pos = str(self.pos)
        self.pos = self.pos.split()
        self.pos = self.pos[1]
        self.pos = self.pos.replace('>','')
        self.pos = self.pos.replace('"','')
        self.pos = self.pos.replace('.',' ')
        self.pos = self.pos.split()
        self.pos = self.pos[1]

    def posPL(self):
        self.posPA = str(self.posPA)
        self.posPA = self.posPA.split()
        self.posPA = self.posPA[1]
        self.posPA = self.posPA.replace('>','')
        self.posPA = self.posPA.replace('"','')
        self.posPA = self.posPA.replace('.',' ')
        self.posPA = self.posPA.split()
        self.posPA = self.posPA[1]

    def ins(self):
        self.ws[self.pos] = float(self.valort)
        self.ws[self.posPA] = float(self.valorp)

    def cons(self):
        a = open(r"txt//clb.txt", "r")
        #LER A PA DA LOJA TOTAL
        def loja():
            cont = 0
            with open(r"txt//clb.txt") as dados:
                linha = dados.readlines()
            for i in linha:
                cont += 1
                s = 3
                if cont == (len(linha) - s):
                    valorLoja = i
                    if valorLoja.count('.') == 1:
                        valorLoja = valorLoja.replace('.', '')
                    valorLoja = valorLoja.replace(',', '.')
                    self.ws['C29'] = float(valorLoja)
                    #self.planilha.save('parcial.xlsx')
                    continue
                if cont == (len(linha) - (s - 1)):
                    valorPa = i
                    if valorPa.count('.') == 1:
                        valorPa = valorPa.replace('.', '')
                    valorPa = valorPa.replace(',', '.')
                    self.ws['E29'] = float(valorPa)
                    #self.planilha.save('parcial.xlsx')
        def ref():
            cont = 3
            for l in range(1, cont):
                self.operador = b
                self.operador = self.operador.split()
                self.nome = str(self.operador[0])
                valort = a.readline()
                if valort.count(".") == 1:
                    valort = valort.replace(".", "")
                self.valort = valort.replace(",", ".")

                valorp = a.readline()
                if valorp.count(".") == 1:
                    valorp = valorp.replace(".", "")
                self.valorp = valorp.replace(",", ".")
                self.pa = a.readline()
                break
        
        for i in range(1, 800):
            b = a.readline()
            for i in range(4,23):
                self.nomeP = self.ws[f'A{i}'].value
                self.nomeP = str(self.nomeP)
                self.pos = self.ws[f'C{i}']
                self.posPA = self.ws[f'E{i}']
                if self.nomeP != None:
                    if self.nomeP in b:
                        self.posP()
                        self.posPL()
                        ref()
                        self.ins()
        loja()
        self.planilha.save(r'pln//parcial.xlsx')
        self.planilha.close()
        os.system('cd pln && parcial.xlsx')
